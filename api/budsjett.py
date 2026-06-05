from http.server import BaseHTTPRequestHandler
import json, base64, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

MRK    = "FF1F4E2E"
MRK2   = "FF163D23"
LYS    = "FFE8F0EA"
LYS2   = "FFD0E4D5"
GUL    = "FFC9A84C"
GUL_LYS= "FFFFF8E8"
ROD_T  = "FF8B2020"
ROD_LYS= "FFFCE8E8"
GRN_T  = "FF1F4E2E"
GRN_LYS= "FFE8F5EC"
CREAM  = "FFF5F0E8"
CREAM2 = "FFEDE7D9"
HVIT   = "FFFFFFFF"
MRK_T  = "FF0F1A12"
GRAA   = "FF5A6E5E"
BLA    = "FFD6E4FF"
BLA_T  = "FF1A3A8B"
GUL_T  = "FF7A5A1E"

def fyll(hex): return PatternFill("solid", start_color=hex, end_color=hex)
def f(bold=False, color="FF000000", size=10, italic=False, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)
def sentr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def venst(): return Alignment(horizontal="left", vertical="center")
def hoyr(): return Alignment(horizontal="right", vertical="center")
def kant():
    t = Side(style="thin", color="FFCCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)
def tykk_kant():
    t = Side(style="medium", color=MRK)
    return Border(left=t, right=t, top=t, bottom=t)

def sett_bredde(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def sett_hoyde(ws, row, h): ws.row_dimensions[row].height = h

KR_FMT = '#,##0" kr";[Red]-#,##0" kr";"-"'
KR_POS = '#,##0" kr"'

def celle(ws, r, c, v=None, bold=False, farge=MRK_T, bg=HVIT, fmt=None, aln="left", size=10, italic=False, wrap=False):
    cel = ws.cell(row=r, column=c, value=v)
    cel.font = f(bold=bold, color=farge, size=size, italic=italic)
    cel.fill = fyll(bg)
    if aln == "right": cel.alignment = hoyr()
    elif aln == "center": cel.alignment = sentr(wrap)
    else: cel.alignment = venst()
    cel.border = kant()
    if fmt: cel.number_format = fmt
    return cel

def tittel_rad(ws, r, c1, c2, tekst, bg=MRK, farge=HVIT, size=14):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cel = ws.cell(row=r, column=c1, value=tekst)
    cel.font = f(bold=True, color=farge, size=size, name="Arial")
    cel.fill = fyll(bg)
    cel.alignment = sentr()
    cel.border = tykk_kant()
    return cel

def seksjon_hdr(ws, r, c1, c2, tekst, bg=MRK, farge=HVIT):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cel = ws.cell(row=r, column=c1, value=tekst)
    cel.font = f(bold=True, color=farge, size=11)
    cel.fill = fyll(bg)
    cel.alignment = sentr()
    cel.border = tykk_kant()
    return cel

def kol_hdr(ws, r, c, tekst, bg=MRK2, farge=HVIT):
    cel = ws.cell(row=r, column=c, value=tekst)
    cel.font = f(bold=True, color=farge, size=9)
    cel.fill = fyll(bg)
    cel.alignment = sentr(True)
    cel.border = kant()
    return cel

def lag_excel(kalkulator, tall):
    wb = Workbook()

    inntekter  = tall.get("inntektLinjer", [])
    kostnader  = tall.get("kostnadLinjer", [])
    skatt_sats = tall.get("skattSats", 0.22)
    inntekt    = tall.get("inntekt", 0)
    total_kost = tall.get("totalKost", 0)
    brutto     = inntekt - total_kost
    skatt      = max(0, brutto * skatt_sats)
    netto      = brutto - skatt

    kalkulator_navn = {
        "eiendom-privat": "EIENDOM PRIVAT",
        "eiendom-as": "EIENDOM VIA AS",
        "bil": "BILUTLEIE",
        "salong": "SALONG"
    }.get(kalkulator, kalkulator.upper())

    # ARK 1: MÅNEDLIG BUDSJETT
    ws = wb.active
    ws.title = "Maanedlig budsjett"
    ws.sheet_view.showGridLines = False

    sett_bredde(ws, 1, 32)
    sett_bredde(ws, 2, 18)
    sett_bredde(ws, 3, 18)
    sett_bredde(ws, 4, 18)
    sett_bredde(ws, 5, 22)

    sett_hoyde(ws, 1, 42)
    tittel_rad(ws, 1, 1, 5, "BUDSJETTARK  -  " + kalkulator_navn, size=16)

    sett_hoyde(ws, 2, 20)
    ws.merge_cells("A2:E2")
    cel = ws["A2"]
    cel.value = "Tall fra kalkulatoren  -  Fyll inn faktiske tall i de bla cellene  -  Avvik beregnes automatisk"
    cel.font = f(color=GRAA, size=9, italic=True)
    cel.alignment = sentr()
    cel.fill = fyll(CREAM)
    cel.border = kant()

    sett_hoyde(ws, 3, 24)
    celle(ws, 3, 1, "GJELDENDE MANED:", bold=True, farge=MRK_T, bg=LYS, size=9)
    cel = ws.cell(row=3, column=2, value="Januar 2026")
    cel.font = f(bold=True, color=GUL_T, size=11)
    cel.fill = fyll(GUL_LYS)
    cel.alignment = venst()
    cel.border = tykk_kant()
    ws.merge_cells("B3:C3")
    celle(ws, 3, 4, "<- Endre maned her", italic=True, farge=GRAA, bg=LYS, size=8)
    celle(ws, 3, 5, None, bg=LYS)

    sett_hoyde(ws, 4, 8)
    for c in range(1, 6): ws.cell(row=4, column=c).fill = fyll(HVIT)

    sett_hoyde(ws, 5, 26)
    seksjon_hdr(ws, 5, 1, 5, "SAMMENDRAG")
    sett_hoyde(ws, 6, 18)
    kol_hdr(ws, 6, 1, "BESKRIVELSE")
    kol_hdr(ws, 6, 2, "BUDSJETT")
    kol_hdr(ws, 6, 3, "FAKTISK")
    kol_hdr(ws, 6, 4, "AVVIK")
    kol_hdr(ws, 6, 5, "AVVIK %")

    samm = [
        ("Inntekter", inntekt, GRN_T, HVIT),
        ("Kostnader", total_kost, ROD_T, HVIT),
        ("Brutto resultat", brutto, GRN_T if brutto>=0 else ROD_T, LYS if brutto>=0 else ROD_LYS),
        ("Skatt (%d%%)" % int(skatt_sats*100), skatt, GUL_T, HVIT),
        ("Netto resultat", netto, GRN_T if netto>=0 else ROD_T, LYS if netto>=0 else ROD_LYS),
    ]
    for i, (navn, bud, farve, bg) in enumerate(samm):
        r = 7 + i
        sett_hoyde(ws, r, 20)
        er_siste = i == len(samm) - 1
        celle(ws, r, 1, navn, bold=er_siste, bg=bg)
        celle(ws, r, 2, bud, bold=er_siste, farge=farve, bg=bg, fmt=KR_FMT, aln="right")
        cel = ws.cell(row=r, column=3, value=None)
        cel.font = f(color=BLA_T, size=10, bold=er_siste)
        cel.fill = fyll(BLA)
        cel.alignment = hoyr()
        cel.border = tykk_kant()
        cel.number_format = KR_FMT
        celle(ws, r, 4, "=IF(C%d=\"\",\"-\",C%d-B%d)" % (r,r,r), bold=er_siste, bg=bg, fmt=KR_FMT, aln="right")
        celle(ws, r, 5, "=IF(OR(C%d=\"\",B%d=0),\"-\",(C%d-B%d)/ABS(B%d))" % (r,r,r,r,r), bold=er_siste, bg=bg, fmt='0.0%;[Red]-0.0%;"-"', aln="right")

    rad = 12

    sett_hoyde(ws, rad, 8)
    for c in range(1, 6): ws.cell(row=rad, column=c).fill = fyll(HVIT)
    rad += 1

    sett_hoyde(ws, rad, 26)
    seksjon_hdr(ws, rad, 1, 5, "INNTEKTER", bg="FF2A6640")
    rad += 1
    sett_hoyde(ws, rad, 18)
    kol_hdr(ws, rad, 1, "BESKRIVELSE")
    kol_hdr(ws, rad, 2, "BUDSJETT")
    kol_hdr(ws, rad, 3, "FAKTISK")
    kol_hdr(ws, rad, 4, "AVVIK")
    kol_hdr(ws, rad, 5, "AVVIK %")
    rad += 1

    inn_start = rad
    for i, l in enumerate(inntekter):
        r = inn_start + i
        sett_hoyde(ws, r, 20)
        bg = CREAM if i % 2 == 0 else HVIT
        celle(ws, r, 1, l["navn"], bg=bg)
        celle(ws, r, 2, l["verdi"], farge=GRN_T, bg=bg, fmt=KR_POS, aln="right")
        cel = ws.cell(row=r, column=3, value=None)
        cel.font = f(color=BLA_T, size=10)
        cel.fill = fyll(BLA)
        cel.alignment = hoyr()
        cel.border = tykk_kant()
        cel.number_format = KR_FMT
        celle(ws, r, 4, "=IF(C%d=\"\",\"-\",C%d-B%d)" % (r,r,r), bg=bg, fmt=KR_FMT, aln="right")
        celle(ws, r, 5, "=IF(OR(C%d=\"\",B%d=0),\"-\",(C%d-B%d)/ABS(B%d))" % (r,r,r,r,r), bg=bg, fmt='0.0%;[Red]-0.0%;"-"', aln="right")
        rad += 1
    inn_slutt = rad - 1

    sett_hoyde(ws, rad, 22)
    celle(ws, rad, 1, "Sum inntekter", bold=True, bg=LYS2)
    celle(ws, rad, 2, "=SUM(B%d:B%d)" % (inn_start, inn_slutt), bold=True, farge=GRN_T, bg=LYS2, fmt=KR_POS, aln="right")
    cel = ws.cell(row=rad, column=3, value="=IF(COUNTA(C%d:C%d)=0,\"-\",SUM(C%d:C%d))" % (inn_start,inn_slutt,inn_start,inn_slutt))
    cel.font = f(bold=True, color=BLA_T, size=10)
    cel.fill = fyll(LYS2)
    cel.alignment = hoyr()
    cel.border = kant()
    cel.number_format = KR_FMT
    celle(ws, rad, 4, "=IF(C%d=\"-\",\"-\",C%d-B%d)" % (rad,rad,rad), bold=True, bg=LYS2, fmt=KR_FMT, aln="right")
    celle(ws, rad, 5, "=IF(OR(C%d=\"-\",B%d=0),\"-\",(C%d-B%d)/ABS(B%d))" % (rad,rad,rad,rad,rad), bold=True, bg=LYS2, fmt='0.0%;[Red]-0.0%;"-"', aln="right")
    rad += 1

    sett_hoyde(ws, rad, 8)
    for c in range(1, 6): ws.cell(row=rad, column=c).fill = fyll(HVIT)
    rad += 1

    sett_hoyde(ws, rad, 26)
    seksjon_hdr(ws, rad, 1, 5, "KOSTNADER", bg="FF6B1A1A")
    rad += 1
    sett_hoyde(ws, rad, 18)
    kol_hdr(ws, rad, 1, "BESKRIVELSE")
    kol_hdr(ws, rad, 2, "BUDSJETT")
    kol_hdr(ws, rad, 3, "FAKTISK")
    kol_hdr(ws, rad, 4, "AVVIK")
    kol_hdr(ws, rad, 5, "AVVIK %")
    rad += 1

    kost_start = rad
    for i, l in enumerate(kostnader):
        r = kost_start + i
        sett_hoyde(ws, r, 20)
        bg = CREAM if i % 2 == 0 else HVIT
        celle(ws, r, 1, l["navn"], bg=bg)
        celle(ws, r, 2, l["verdi"], farge=ROD_T, bg=bg, fmt=KR_POS, aln="right")
        cel = ws.cell(row=r, column=3, value=None)
        cel.font = f(color=BLA_T, size=10)
        cel.fill = fyll(BLA)
        cel.alignment = hoyr()
        cel.border = tykk_kant()
        cel.number_format = KR_FMT
        celle(ws, r, 4, "=IF(C%d=\"\",\"-\",C%d-B%d)" % (r,r,r), bg=bg, fmt=KR_FMT, aln="right")
        celle(ws, r, 5, "=IF(OR(C%d=\"\",B%d=0),\"-\",(C%d-B%d)/ABS(B%d))" % (r,r,r,r,r), bg=bg, fmt='0.0%;[Red]-0.0%;"-"', aln="right")
        rad += 1
    kost_slutt = rad - 1

    sett_hoyde(ws, rad, 22)
    celle(ws, rad, 1, "Sum kostnader", bold=True, bg=ROD_LYS)
    celle(ws, rad, 2, "=SUM(B%d:B%d)" % (kost_start, kost_slutt), bold=True, farge=ROD_T, bg=ROD_LYS, fmt=KR_POS, aln="right")
    cel = ws.cell(row=rad, column=3, value="=IF(COUNTA(C%d:C%d)=0,\"-\",SUM(C%d:C%d))" % (kost_start,kost_slutt,kost_start,kost_slutt))
    cel.font = f(bold=True, color=BLA_T, size=10)
    cel.fill = fyll(ROD_LYS)
    cel.alignment = hoyr()
    cel.border = kant()
    cel.number_format = KR_FMT
    celle(ws, rad, 4, "=IF(C%d=\"-\",\"-\",C%d-B%d)" % (rad,rad,rad), bold=True, bg=ROD_LYS, fmt=KR_FMT, aln="right")
    celle(ws, rad, 5, "=IF(OR(C%d=\"-\",B%d=0),\"-\",(C%d-B%d)/ABS(B%d))" % (rad,rad,rad,rad,rad), bold=True, bg=ROD_LYS, fmt='0.0%;[Red]-0.0%;"-"', aln="right")
    rad += 1

    sett_hoyde(ws, rad, 8)
    for c in range(1, 6): ws.cell(row=rad, column=c).fill = fyll(HVIT)
    rad += 1

    bg_netto = MRK if netto >= 0 else "FF6B1A1A"
    sett_hoyde(ws, rad, 26)
    seksjon_hdr(ws, rad, 1, 5, "NETTO RESULTAT PER MANED", bg=bg_netto)
    rad += 1
    sett_hoyde(ws, rad, 32)
    bg_r = LYS if netto >= 0 else ROD_LYS
    celle(ws, rad, 1, "Netto per maned (etter skatt)", bold=True, bg=bg_r, size=11)
    celle(ws, rad, 2, netto, bold=True, farge=GRN_T if netto>=0 else ROD_T, bg=bg_r, fmt=KR_FMT, aln="right", size=13)
    cel = ws.cell(row=rad, column=3, value=None)
    cel.font = f(bold=True, color=BLA_T, size=13)
    cel.fill = fyll(BLA)
    cel.alignment = hoyr()
    cel.border = tykk_kant()
    cel.number_format = KR_FMT
    celle(ws, rad, 4, "=IF(C%d=\"\",\"-\",C%d-B%d)" % (rad,rad,rad), bold=True, bg=bg_r, fmt=KR_FMT, aln="right", size=11)
    celle(ws, rad, 5, "=IF(OR(C%d=\"-\",B%d=0),\"-\",(C%d-B%d)/ABS(B%d))" % (rad,rad,rad,rad,rad), bold=True, bg=bg_r, fmt='0.0%;[Red]-0.0%;"-"', aln="right")
    netto_rad = rad
    rad += 1

    sett_hoyde(ws, rad, 22)
    celle(ws, rad, 1, "Estimert arsresultat", bg=bg_r)
    celle(ws, rad, 2, "=B%d*12" % netto_rad, farge=GRN_T if netto>=0 else ROD_T, bg=bg_r, fmt=KR_FMT, aln="right")
    celle(ws, rad, 3, "=IF(C%d=\"\",\"-\",C%d*12)" % (netto_rad, netto_rad), bg=BLA, fmt=KR_FMT, aln="right")
    celle(ws, rad, 4, "=IF(C%d=\"-\",\"-\",C%d-B%d)" % (rad,rad,rad), bg=bg_r, fmt=KR_FMT, aln="right")
    celle(ws, rad, 5, "=IF(OR(C%d=\"-\",B%d=0),\"-\",(C%d-B%d)/ABS(B%d))" % (rad,rad,rad,rad,rad), bg=bg_r, fmt='0.0%;[Red]-0.0%;"-"', aln="right")
    rad += 2

    ws.freeze_panes = "A7"

    # Kakediagram
    diag_start = rad
    celle(ws, diag_start, 1, "KOSTNADSFORDELING", bold=True, bg=LYS, farge=MRK_T)
    ws.merge_cells(start_row=diag_start, start_column=1, end_row=diag_start, end_column=2)
    for i, l in enumerate(kostnader):
        r_d = diag_start + 1 + i
        sett_hoyde(ws, r_d, 16)
        celle(ws, r_d, 1, l["navn"], bg=CREAM if i%2==0 else HVIT, size=9)
        celle(ws, r_d, 2, l["verdi"], bg=CREAM if i%2==0 else HVIT, fmt=KR_POS, aln="right", size=9)

    pie = PieChart()
    pie.title = "Kostnadsfordeling"
    pie.style = 26
    pie.width = 14
    pie.height = 12
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=diag_start+1, max_row=diag_start+len(kostnader))
    cats = Reference(ws, min_col=1, max_col=1, min_row=diag_start+1, max_row=diag_start+len(kostnader))
    pie.add_data(data_ref)
    pie.set_categories(cats)
    ws.add_chart(pie, "D%d" % diag_start)

    # ARK 2: 12-MÅNEDER
    ws2 = wb.create_sheet("12-maneder")
    ws2.sheet_view.showGridLines = False

    sett_bredde(ws2, 1, 10)
    for c in range(2, 10): sett_bredde(ws2, c, 16)

    sett_hoyde(ws2, 1, 42)
    tittel_rad(ws2, 1, 1, 9, "12-MANEDERS OVERSIKT  -  " + kalkulator_navn, size=15)

    sett_hoyde(ws2, 2, 20)
    ws2.merge_cells("A2:I2")
    cel = ws2["A2"]
    cel.value = "Automatisk beregning  -  Fyll inn faktiske nettotall i H-kolonnen (bla) for avviksanalyse"
    cel.font = f(color=GRAA, size=9, italic=True)
    cel.alignment = sentr()
    cel.fill = fyll(CREAM)
    cel.border = kant()

    sett_hoyde(ws2, 3, 8)
    for c in range(1, 10): ws2.cell(row=3, column=c).fill = fyll(HVIT)

    sett_hoyde(ws2, 4, 22)
    headers2 = ["MANED","INNTEKT","KOSTNADER","BRUTTO","SKATT","NETTO","AKKUMULERT","FAKTISK NETTO","AVVIK"]
    bg_hdr = [MRK2, "FF2A6640", "FF6B1A1A", MRK2, "FF5A4010", MRK2, MRK2, BLA_T, MRK2]
    for c, (h, bg) in enumerate(zip(headers2, bg_hdr), 1):
        cel = ws2.cell(row=4, column=c, value=h)
        cel.font = f(bold=True, color=HVIT if c != 8 else BLA_T, size=9)
        cel.fill = fyll(bg if c != 8 else BLA)
        cel.alignment = sentr(True)
        cel.border = kant()

    maaneder = ["Jan","Feb","Mar","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Des"]
    for i, mnd in enumerate(maaneder):
        r = 5 + i
        sett_hoyde(ws2, r, 20)
        bg = CREAM if i % 2 == 0 else HVIT
        celle(ws2, r, 1, mnd, bold=True, bg=bg)
        celle(ws2, r, 2, inntekt, farge=GRN_T, bg=bg, fmt=KR_POS, aln="right")
        celle(ws2, r, 3, total_kost, farge=ROD_T, bg=bg, fmt=KR_POS, aln="right")
        celle(ws2, r, 4, "=B%d-C%d" % (r,r), bg=bg, fmt=KR_FMT, aln="right")
        celle(ws2, r, 5, "=MAX(0,D%d*%s)" % (r, skatt_sats), farge=GUL_T, bg=bg, fmt=KR_POS, aln="right")
        celle(ws2, r, 6, "=D%d-E%d" % (r,r), bold=True, bg=bg, fmt=KR_FMT, aln="right")
        akk = "=F%d" % r if i == 0 else "=G%d+F%d" % (r-1, r)
        celle(ws2, r, 7, akk, bold=True, bg=bg, fmt=KR_FMT, aln="right")
        cel = ws2.cell(row=r, column=8, value=None)
        cel.font = f(color=BLA_T, size=10)
        cel.fill = fyll(BLA)
        cel.alignment = hoyr()
        cel.border = tykk_kant()
        cel.number_format = KR_FMT
        celle(ws2, r, 9, "=IF(H%d=\"\",\"-\",H%d-F%d)" % (r,r,r), bg=bg, fmt=KR_FMT, aln="right")

    sett_hoyde(ws2, 17, 26)
    celle(ws2, 17, 1, "TOTALT", bold=True, farge=HVIT, bg=MRK, aln="center")
    totals = ["=SUM(B5:B16)","=SUM(C5:C16)","=SUM(D5:D16)","=SUM(E5:E16)","=SUM(F5:F16)","=G16",
              "=IF(COUNTA(H5:H16)=0,\"-\",SUM(H5:H16))","=IF(H17=\"-\",\"-\",H17-F17)"]
    for c, frm in enumerate(totals, 2):
        celle(ws2, 17, c, frm, bold=True, farge=HVIT, bg=MRK, fmt=KR_FMT, aln="right")

    ws2.freeze_panes = "A5"

    # Søylediagram
    sett_hoyde(ws2, 19, 16)
    celle(ws2, 19, 1, "NETTO PER MANED - BUDSJETT vs FAKTISK", bold=True, bg=LYS, farge=MRK_T)
    ws2.merge_cells("A19:I19")

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Netto per maned"
    bar.y_axis.title = "NOK"
    bar.style = 10
    bar.width = 32
    bar.height = 14
    bud_ref = Reference(ws2, min_col=6, max_col=6, min_row=4, max_row=16)
    fak_ref = Reference(ws2, min_col=8, max_col=8, min_row=4, max_row=16)
    cats2 = Reference(ws2, min_col=1, min_row=5, max_row=16)
    bar.add_data(bud_ref, titles_from_data=True)
    bar.add_data(fak_ref, titles_from_data=True)
    bar.set_categories(cats2)
    bar.series[0].graphicalProperties.solidFill = "1F4E2E"
    bar.series[1].graphicalProperties.solidFill = "4A9EDB"
    ws2.add_chart(bar, "A20")

    # Linjediagram akkumulert
    line = LineChart()
    line.title = "Akkumulert netto"
    line.style = 10
    line.y_axis.title = "NOK"
    line.width = 32
    line.height = 14
    akk_ref = Reference(ws2, min_col=7, max_col=7, min_row=4, max_row=16)
    line.add_data(akk_ref, titles_from_data=True)
    line.set_categories(cats2)
    line.series[0].graphicalProperties.line.solidFill = "C9A84C"
    line.series[0].graphicalProperties.line.width = 28000
    ws2.add_chart(line, "A36")

    # ARK 3: NOKKELTALL
    ws3 = wb.create_sheet("Nokkeltall")
    ws3.sheet_view.showGridLines = False
    sett_bredde(ws3, 1, 32)
    sett_bredde(ws3, 2, 20)
    sett_bredde(ws3, 3, 22)

    sett_hoyde(ws3, 1, 42)
    tittel_rad(ws3, 1, 1, 3, "NOKKELTALL  -  " + kalkulator_navn, size=15)

    sett_hoyde(ws3, 3, 18)
    kol_hdr(ws3, 3, 1, "NOKKELTAL")
    kol_hdr(ws3, 3, 2, "VERDI")
    kol_hdr(ws3, 3, 3, "STATUS")

    metrics = [
        ("Manedlig inntekt (budsjett)", inntekt, GRN_T, LYS, KR_FMT),
        ("Manedlige kostnader (budsjett)", total_kost, ROD_T, ROD_LYS, KR_FMT),
        ("Brutto resultat/mnd", brutto, GRN_T if brutto>=0 else ROD_T, LYS if brutto>=0 else ROD_LYS, KR_FMT),
        ("Skatt per maned", skatt, GUL_T, GUL_LYS, KR_FMT),
        ("Netto per maned", netto, GRN_T if netto>=0 else ROD_T, LYS if netto>=0 else ROD_LYS, KR_FMT),
        ("Estimert arsresultat", netto*12, GRN_T if netto>=0 else ROD_T, LYS if netto>=0 else ROD_LYS, KR_FMT),
        ("Kostnadsprosent", total_kost/inntekt if inntekt>0 else 0, ROD_T, CREAM, '0.0%'),
        ("Nettomarginn", netto/inntekt if inntekt>0 else 0, GRN_T if netto>=0 else ROD_T, CREAM, '0.0%'),
    ]
    statuser = ["", "", "", "", "Lonnsomt" if netto>=0 else "Negativt", "", "God" if (inntekt>0 and total_kost/inntekt<0.7) else "Hoy", ""]
    for i, (navn, verdi, farve, bg, fmt) in enumerate(metrics):
        r = 4 + i
        sett_hoyde(ws3, r, 22)
        celle(ws3, r, 1, navn, bg=bg)
        celle(ws3, r, 2, verdi, bold=True, farge=farve, bg=bg, fmt=fmt, aln="right")
        celle(ws3, r, 3, statuser[i], bold=bool(statuser[i]), farge=GRN_T if "Lonnsomt" in statuser[i] or "God" in statuser[i] else ROD_T if statuser[i] else MRK_T, bg=bg)

    # ARK 4: BRUKERGUIDE
    ws4 = wb.create_sheet("Brukerguide")
    ws4.sheet_view.showGridLines = False
    sett_bredde(ws4, 1, 14)
    sett_bredde(ws4, 2, 55)

    sett_hoyde(ws4, 1, 42)
    tittel_rad(ws4, 1, 1, 2, "BRUKERGUIDE  -  INVEST TOOLS BY ADDON", size=14)

    guide = [
        ("FARGEKODE", "BETYR", True),
        ("Bla celle", "Fyll inn faktiske tall her - disse cellene er interaktive", False),
        ("Hvit/kremfarget celle", "Budsjett-tall fra kalkulatoren - ikke endre disse", False),
        ("Gronn tekst", "Inntekter eller positiv verdi", False),
        ("Rod tekst", "Kostnader eller negativ verdi", False),
        (None, None, False),
        ("AVVIK", "BETYR", True),
        ("Positivt tall", "Bedre enn budsjett - du tjente mer eller brukte mindre", False),
        ("Negativt tall", "Verre enn budsjett - lavere inntekt eller hoyere kostnader", False),
        ("-", "Faktisk tall ikke fylt inn ennaa", False),
        (None, None, False),
        ("TIPS", "ANBEFALING", True),
        ("Manedlig rutine", "Fyll inn faktiske tall ved manedsslutt for best oversikt", False),
        ("12-maneder ark", "Se hele arets utvikling og sammenlign budsjett vs faktisk", False),
        ("Nokkeltall ark", "Fa rask oversikt over lonnsamhet og nokkelrater", False),
        (None, None, False),
        ("ADVARSEL", "Tallene er estimater og ikke finansiell radgivning. Konsulter en regnskapsforer.", False),
    ]
    bg_map = {"FARGEKODE": MRK2, "AVVIK": MRK2, "TIPS": MRK2, "ADVARSEL": ROD_LYS}
    fg_map = {"ADVARSEL": ROD_T}
    for i, (k, v, is_hdr) in enumerate(guide):
        r = 2 + i
        sett_hoyde(ws4, r, 20)
        if k is None:
            for c in [1, 2]:
                ws4.cell(row=r, column=c).fill = fyll(HVIT)
                ws4.cell(row=r, column=c).border = kant()
            continue
        bg = bg_map.get(k, CREAM if i%2==0 else HVIT)
        fg = fg_map.get(k, HVIT if is_hdr else MRK_T)
        for c, txt in enumerate([k, v], 1):
            cel = ws4.cell(row=r, column=c, value=txt)
            cel.font = f(bold=is_hdr, color=fg, size=10)
            cel.fill = fyll(bg)
            cel.alignment = venst()
            cel.border = kant()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        data = json.loads(self.rfile.read(length))
        result = lag_excel(data.get("kalkulator", "eiendom-privat"), data.get("tall", {}))
        body = json.dumps({"xlsx": result}).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()